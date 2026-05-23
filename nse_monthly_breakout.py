import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
import json
import requests
import warnings
warnings.filterwarnings("ignore")

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
INPUT_CSV   = os.path.join(BASE_DIR, "data", "nse500list.csv")
OUTPUT_DIR  = os.path.join(BASE_DIR, "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Monthly_Breakout_Report.xlsx")
SUMMARY_FILE= os.path.join(OUTPUT_DIR, "summary.json")   # read by workflow step

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Telegram ──────────────────────────────────────────────────────────────────
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID   = os.environ.get("TELEGRAM_CHAT_ID", "")
PAGES_URL          = os.environ.get("PAGES_URL", "")
# Set SEND_TELEGRAM=true only in the dedicated workflow step (after Pages deploy)
SEND_TELEGRAM      = os.environ.get("SEND_TELEGRAM", "false").lower() == "true"


def send_telegram(msg: str) -> None:
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("[Telegram] Credentials not set — skipping.")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": msg, "parse_mode": "HTML"}
    try:
        r = requests.post(url, json=payload, timeout=10)
        print("[Telegram] Sent ✓" if r.ok else f"[Telegram] Failed: {r.status_code} {r.text}")
    except Exception as e:
        print(f"[Telegram] Exception: {e}")


# ── Helpers ───────────────────────────────────────────────────────────────────
def load_symbols(csv_path: str) -> list:
    df = pd.read_csv(csv_path)
    for col in df.columns:
        if col.strip().lower() in ("symbol", "ticker", "scrip", "nsesymbol", "nse symbol"):
            syms = df[col].dropna().str.strip().tolist()
            return [s + ".NS" if not s.endswith(".NS") else s for s in syms]
    syms = df.iloc[:, 0].dropna().str.strip().tolist()
    return [s + ".NS" if not s.endswith(".NS") else s for s in syms]


def get_monthly_data(symbol: str):
    try:
        tk = yf.Ticker(symbol)
        df = tk.history(period="6mo", interval="1mo", auto_adjust=True)
        if df.empty or len(df) < 2:
            return None
        today = date.today()
        last_bar_date = df.index[-1].date()
        if last_bar_date.year == today.year and last_bar_date.month == today.month:
            df = df.iloc[:-1]
        if len(df) < 2:
            return None
        prev  = df.iloc[-2]
        recnt = df.iloc[-1]
        return {
            "prev_high"  : round(prev["High"],  2),
            "prev_low"   : round(prev["Low"],   2),
            "prev_close" : round(prev["Close"], 2),
            "prev_month" : df.index[-2].strftime("%b %Y"),
            "rec_high"   : round(recnt["High"],  2),
            "rec_low"    : round(recnt["Low"],   2),
            "rec_close"  : round(recnt["Close"], 2),
            "rec_month"  : df.index[-1].strftime("%b %Y"),
        }
    except Exception:
        return None


# ── Fetch & filter ────────────────────────────────────────────────────────────
symbols = load_symbols(INPUT_CSV)
print(f"Loaded {len(symbols)} symbols.")

results = []
for i, sym in enumerate(symbols, 1):
    print(f"[{i:>4}/{len(symbols)}] {sym}", end="\r")
    data = get_monthly_data(sym)
    if data is None:
        continue
    if data["rec_close"] > data["prev_high"]:
        double_confirm = "YES" if data["rec_low"] > data["prev_low"] else ""
        results.append({
            "Ticker"               : sym.replace(".NS", ""),
            "Prev Month"           : data["prev_month"],
            "Prev Month High"      : data["prev_high"],
            "Prev Month Low"       : data["prev_low"],
            "Recent Month"         : data["rec_month"],
            "Recent Monthly Close" : data["rec_close"],
            "Recent Monthly Low"   : data["rec_low"],
            "Close > Prev High"    : "✓",
            "Low > Prev Low (YES)" : double_confirm,
        })

full_breakout = sum(1 for r in results if r["Low > Prev Low (YES)"] == "YES")
run_month     = results[0]["Recent Month"] if results else datetime.now().strftime("%b %Y")
print(f"\nFound {len(results)} breakout stocks  |  Full breakout: {full_breakout}")


# ── Build Excel ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Monthly Breakout"

HDR_FILL  = PatternFill("solid", fgColor="1F3864")
YES_FILL  = PatternFill("solid", fgColor="00B050")
TICK_FILL = PatternFill("solid", fgColor="D9EAD3")
ALT_FILL  = PatternFill("solid", fgColor="EEF2FF")
WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
thin = Side(style="thin", color="BFBFBF")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(name="Arial", size=10)
YES_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=13, color="1F3864")

ws.merge_cells("A1:I1")
ws["A1"] = f"NSE Monthly Breakout Screener  —  Generated: {datetime.now().strftime('%d %b %Y %H:%M')} IST"
ws["A1"].font      = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

headers    = ["Ticker","Prev Month","Prev Month\nHigh (₹)","Prev Month\nLow (₹)",
              "Recent Month","Recent Monthly\nClose (₹)","Recent Monthly\nLow (₹)",
              "Close >\nPrev High","Low > Prev Low\n(Full Breakout)"]
col_widths = [14, 12, 16, 16, 13, 18, 18, 13, 18]

ws.row_dimensions[2].height = 36
for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=2, column=col_idx, value=hdr)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = bdr
    ws.column_dimensions[get_column_letter(col_idx)].width = w

for row_idx, rec in enumerate(results, 3):
    row_fill = ALT_FILL if row_idx % 2 == 0 else WHT_FILL
    values = [rec["Ticker"], rec["Prev Month"], rec["Prev Month High"], rec["Prev Month Low"],
              rec["Recent Month"], rec["Recent Monthly Close"], rec["Recent Monthly Low"],
              rec["Close > Prev High"], rec["Low > Prev Low (YES)"]]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = BODY_FONT
        cell.border    = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in (3, 4, 6, 7) and isinstance(val, (int, float)):
            cell.number_format = "#,##0.00"
        if col_idx == 9 and val == "YES":
            cell.fill = YES_FILL
            cell.font = YES_FONT
        elif col_idx == 8 and val == "✓":
            cell.fill = TICK_FILL
        else:
            cell.fill = row_fill
    ws.row_dimensions[row_idx].height = 18

sum_row = 2 + len(results) + 2
ws.merge_cells(f"A{sum_row}:D{sum_row}")
ws[f"A{sum_row}"] = f"Total Breakout Stocks: {len(results)}"
ws[f"A{sum_row}"].font = Font(name="Arial", bold=True, size=10, color="1F3864")
ws.merge_cells(f"A{sum_row+1}:D{sum_row+1}")
ws[f"A{sum_row+1}"] = f"Full Breakout (Close > Prev High  AND  Low > Prev Low): {full_breakout}"
ws[f"A{sum_row+1}"].font = Font(name="Arial", bold=True, size=10, color="375623")
ws.freeze_panes = "A3"

wb.save(OUTPUT_FILE)
print(f"Report saved → {OUTPUT_FILE}")

# ── Save summary.json (read by the Telegram workflow step) ────────────────────
top_tickers = [r["Ticker"] for r in results if r["Low > Prev Low (YES)"] == "YES"][:10]
summary = {
    "run_month"     : run_month,
    "breakout_count": len(results),
    "full_breakout" : full_breakout,
    "top_tickers"   : top_tickers,
}
with open(SUMMARY_FILE, "w") as f:
    json.dump(summary, f)
print(f"Summary saved → {SUMMARY_FILE}")

# ── Telegram (only when SEND_TELEGRAM=true, i.e. after Pages is live) ─────────
if SEND_TELEGRAM:
    top_line = ("  " + "  |  ".join(summary["top_tickers"])) if summary["top_tickers"] else "  —"
    dl_line  = f'\n📥 <a href="{PAGES_URL}">Download Report (Excel)</a>' if PAGES_URL else ""
    msg = (
        f"📊 <b>NSE Monthly Breakout — {run_month}</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"✅ Breakout stocks  : <b>{len(results)}</b>\n"
        f"🔥 Full breakout    : <b>{full_breakout}</b>  (Close &amp; Low both above prev month)\n"
        f"\n<b>Top Full-Breakout Tickers:</b>\n{top_line}"
        f"{dl_line}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"<i>Universe: NSE 500  |  Source: Yahoo Finance</i>"
    )
    send_telegram(msg)
